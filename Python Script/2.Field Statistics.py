
# ---------------------------------------------
# 2.Field Statistics.py
#
# 本脚本实现对数据文件的字段级统计分析，包括唯一值、缺失值、常见取值等，并输出Excel报告。
# 主要流程：
#   2.1 读取数据文件。
#   2.2 对数据文件中的每一列进行统计分析，计算唯一值数量、缺失值数量、非缺失值数量、前10个最常见的值及其占比等。
#   2.3 整合数据，将之写入到Excel文件中，便于后续查看和分析。
# ---------------------------------------------


# 2.1 读取数据文件。

# 导入必要的库，其中openpyxl专门用于编辑excel
import pandas as pd
import dask
import dask.dataframe as dd
from openpyxl import Workbook

# 数据文件目录
file_path = r'.\data'

# 读取销售数据，采用Dask以支持大文件处理
sales_data = dd.read_csv(
    file_path + r'\with_sales.csv',
    encoding='UTF-8',
    encoding_errors='ignore',
    assume_missing=True
)
nosales_data = dd.read_csv(
    file_path+r'\without_sales.csv',
    encoding='UTF-8',
    encoding_errors='ignore',
    assume_missing=True
    )


# 2.2 对数据文件中的每一列进行统计分析，计算唯一值数量、缺失值数量、非缺失值数量、前10个最常见的值及其占比等。

# 针对单列数据进行统计分析，返回各类统计指标
def analyze_column(column, col_name):
    # 计算唯一值数量和缺失值数量
    unique_count = column.nunique().compute()
    null_count = column.isnull().sum().compute()
    total_count = len(column)

    # 计算非缺失值数量
    non_null_count = total_count - null_count

    # 获取非缺失值的值计数（TOP10）
    value_counts = column.value_counts().compute()

    # 获取前10个最常见的值
    top_10 = top_10 = value_counts.sort_values(ascending=False).head(10)

    # 计算前10个值的总行数和占比
    top_10_total = top_10.sum()
    top_10_percentage = top_10_total / non_null_count * 100

    # 计算其他值的行数和占比
    other_total = non_null_count - top_10_total
    other_percentage = 100 - top_10_percentage

    # 汇总结果
    result = {
        'column': col_name,
        'unique_count': unique_count,
        'null_count': null_count,
        'null_percentage': null_count / total_count * 100,
        'top_10_values': top_10.index.tolist(),
        'top_10_counts': top_10.values.tolist(),
        'top_10_percentages': (top_10.values / non_null_count * 100).tolist(),
        'other_total': other_total,
        'other_percentage': other_percentage
    }
    return result

# 分析所有列
results = []
for col in nosales_data.columns:
    print(f"正在分析列: {col}")
    try:
        result = analyze_column(nosales_data[col], col)
        results.append(result)
    except Exception as e:
        print(f"分析列 {col} 时出错: {e}")
        results.append({
            'column': col,
            'error': str(e)
        })

# 将结果转换为 DataFrame 以便更好地查看
nosales_result = pd.DataFrame(results)
print(nosales_result.head())

# 分析所有列
results = []
for col in sales_data.columns:
    print(f"正在分析列: {col}")
    try:
        result = analyze_column(sales_data[col], col)
        results.append(result)
    except Exception as e:
        print(f"分析列 {col} 时出错: {e}")
        results.append({
            'column': col,
            'error': str(e)
        })

# 将结果转换为 DataFrame 以便更好地查看
sales_result = pd.DataFrame(results)
print(sales_result.head())


# 2.3 整合数据，将之写入到Excel文件中，便于后续查看和分析。

# 将统计结果写入Excel文件，每列一个Sheet，便于后续分析
def write_result_to_excel(df, excel_path):
    wb = Workbook()
    # 删除openpyxl默认的Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    for idx, row in df.iterrows():
        # Excel Sheet名称最长31字符
        sheet_name = str(row['column'])[:31] if str(row['column']) else f'Sheet{idx+1}'
        ws = wb.create_sheet(title=sheet_name)
        # 写入字段统计内容
        ws['A1'] = row['column']
        ws['A2'] = '唯一值数量'
        ws['B2'] = row['unique_count']
        ws['A3'] = '空值数量'
        ws['B3'] = row['null_count']
        ws['C3'] = row['null_percentage']
        # Top 10常见取值
        top10_values = row['top_10_values'] if isinstance(row['top_10_values'], list) else []
        top10_counts = row['top_10_counts'] if isinstance(row['top_10_counts'], list) else []
        top10_percentages = row['top_10_percentages'] if isinstance(row['top_10_percentages'], list) else []
        for i in range(10):
            ws[f'A{4+i}'] = top10_values[i] if i < len(top10_values) else ''
            ws[f'B{4+i}'] = top10_counts[i] if i < len(top10_counts) else ''
            ws[f'C{4+i}'] = top10_percentages[i] if i < len(top10_percentages) else ''
        # 其他取值统计
        ws['A15'] = '其他取值'
        ws['B15'] = row['other_total']
        ws['C15'] = row['other_percentage']
    wb.save(excel_path)

# 执行写入Excel操作，输出最终统计结果
write_result_to_excel(sales_result, file_path + r'\sales_TOP10.xlsx')
write_result_to_excel(nosales_result, file_path + r'\nosales_TOP10.xlsx')